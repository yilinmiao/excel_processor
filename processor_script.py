from openpyxl import load_workbook

wb = load_workbook(filename=r'2174 Schedule Planning  - Instructor Analysis.xlsx')
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])

data_dic = {}
for rx in range(1, ws.max_row + 1):
    temp_list = []
    pid = rx - 1
    w1 = ws.cell(row=rx, column=1).value
    w2 = ws.cell(row=rx, column=2).value
    w3 = ws.cell(row=rx, column=3).value
    w4 = ws.cell(row=rx, column=4).value
    w5 = ws.cell(row=rx, column=5).value
    w6 = ws.cell(row=rx, column=6).value
    w7 = ws.cell(row=rx, column=7).value
    temp_list = [w1, w2, w3, w4, w5, w6, w7]
    data_dic[pid] = temp_list

data_dic2 = {}
data_dic2[0] = ['Instructor', 'Classes']
subject = ''
num = 0;
section = 0;
i = 0;
for rx in range(1, ws.max_row):
    if ((data_dic[rx][0] is None) and (data_dic[rx][4] is None)):
        continue;
    if (data_dic[rx][1] is None):
        data_dic[rx][1] = subject
    else:
        subject = data_dic[rx][1]
    if (data_dic[rx][2] is None):
        data_dic[rx][2] = num
    else:
        num = data_dic[rx][2]
    if (data_dic[rx][3] is None):
        data_dic[rx][3] = section
    else:
        section = data_dic[rx][3]
    if (data_dic[rx][3] <= 9):
        data_dic[rx][3] = '0' + str(data_dic[rx][3])
    if (data_dic[rx][4] != 'TBA'):
        if (data_dic[rx][5] < 1000):
            data_dic[rx][5] = (str)(data_dic[rx][5])[0:1] + ':' + (str)(data_dic[rx][5])[1:]
        else:
            data_dic[rx][5] = (str)(data_dic[rx][5])[0:2] + ':' + (str)(data_dic[rx][5])[2:]
        if (data_dic[rx][6] < 1000):
            data_dic[rx][6] = (str)(data_dic[rx][6])[0:1] + ':' + (str)(data_dic[rx][6])[1:]
        else:
            data_dic[rx][6] = (str)(data_dic[rx][6])[0:2] + ':' + (str)(data_dic[rx][6])[2:]
        temp = data_dic[rx][1] + ' ' + (str)(data_dic[rx][2]) + '-' + (str)(data_dic[rx][3]) + ', ' + data_dic[rx][4] + ', ' + (str)(data_dic[rx][5]) + '-' + (str)(data_dic[rx][6])
    else:
        temp = data_dic[rx][1] + ' ' + (str)(data_dic[rx][2]) + '-' + (str)(data_dic[rx][3]) + ', ' + data_dic[rx][4]

    if ((data_dic[rx][0] is None)):
        data_dic2[i].append(temp)
    else:
        i += 1
        data_dic2[i] = [data_dic[rx][0], temp]

# print(data_dic)
# print(data_dic2)

ws3 = wb.get_sheet_by_name(sheetnames[1])
for key in data_dic2:
    for col in range(1, len(data_dic2[key]) + 1):
        ws3.cell(column=col, row=key + 1, value=data_dic2[key][col - 1])
wb.save('empty_book.xlsx')



